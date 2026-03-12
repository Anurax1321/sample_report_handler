#!/usr/bin/env python3
"""Extract all data from the Excel file for analysis."""
import openpyxl
from openpyxl.utils import get_column_letter

FILE = 'test_Samples/18022026 MR LABS DATA.xlsm'
wb = openpyxl.load_workbook(FILE, keep_vba=True, data_only=False)

print('=' * 120)
print(f'FILE: {FILE}')
print(f'Sheet names: {wb.sheetnames}')
print(f'Has VBA/Macros: {wb.vba_archive is not None}')
print('=' * 120)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n{"#" * 120}')
    print(f'SHEET: "{sheet_name}"')
    print(f'Dimensions: {ws.dimensions}')
    print(f'Min row: {ws.min_row}, Max row: {ws.max_row}, Min col: {ws.min_column}, Max col: {ws.max_column}')
    merged = list(ws.merged_cells.ranges) if ws.merged_cells else []
    print(f'Merged cells: {merged if merged else "None"}')
    print(f'{"#" * 120}')

    actual_max_col = ws.max_column
    max_rows = min(ws.max_row, 50)

    print(f'\n--- ALL DATA (rows 1 to {max_rows}, cols 1 to {actual_max_col}) ---')

    for row_idx in range(1, max_rows + 1):
        row_data = []
        for col_idx in range(1, actual_max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if isinstance(val, str) and val.startswith('='):
                row_data.append(f'[FORMULA:{val}]')
            elif val is not None:
                row_data.append(str(val))
            else:
                row_data.append('')
        if any(v != '' for v in row_data):
            print(f'Row {row_idx:3d}: {" | ".join(row_data)}')

    print(f'\n--- Column details ---')
    for col_idx in range(1, actual_max_col + 1):
        header = ws.cell(row=1, column=col_idx).value
        col_letter = get_column_letter(col_idx)
        types_seen = set()
        for row_idx in range(2, min(ws.max_row + 1, 25)):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                types_seen.add(type(cell.value).__name__)
        print(f'  Col {col_letter} ({col_idx}): Header="{header}", DataTypes={types_seen}')

print(f'\n{"=" * 120}')
print('DEFINED NAMES / NAMED RANGES:')
for dn in wb.defined_names.definedName:
    print(f'  {dn.name} -> {dn.attr_text}')

print(f'\n{"=" * 120}')
print('VBA MACRO INFO:')
if wb.vba_archive:
    print(f'  VBA archive contains: {wb.vba_archive.namelist()}')
else:
    print('  No VBA macros found')

wb.close()
