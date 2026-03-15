# Excel data extraction module for NBS laboratory Excel files (.xlsm/.xlsx)
# Handles parsing of Shimadzu NeoBase Excel exports (e.g., "18022026 MR LABS DATA.xlsm")
# Produces the same DataFrame format as the text file pipeline (data_extraction.get_final_data)
#
# Supports two sheet layouts:
#   - "Concentration" sheet (preferred): transposed layout, samples as columns, compounds as rows.
#     Extracts only patient columns (skips BLANK and CONTROL columns).
#   - "ConcData" sheet (fallback): row-based layout, samples as rows.
#     Includes controls (duplicated to 4 rows for downstream pipeline).

import re
import openpyxl
import pandas as pd
import numpy as np
from typing import Tuple, List


class ExcelDataExtractionError(Exception):
    """Custom exception for Excel data extraction errors"""
    pass


def extract_from_excel(file_path: str) -> Tuple[pd.DataFrame, List[str], int]:
    """
    Extract concentration data from a Shimadzu NeoBase Excel file.

    Tries the 'Concentration' sheet first (patients only, no controls).
    Falls back to 'ConcData' sheet (includes controls).

    Args:
        file_path: Path to the .xlsm or .xlsx file

    Returns:
        Tuple of (combined_df, sample_names, sample_count)
        - combined_df: DataFrame with 'Sample text' column + all compound columns
        - sample_names: List of sample names
        - sample_count: Total number of samples

    Raises:
        ExcelDataExtractionError: If file format is invalid or data extraction fails
    """
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
    except Exception as e:
        raise ExcelDataExtractionError(f"Failed to open Excel file: {e}")

    if 'Concentration' in wb.sheetnames:
        return _extract_from_concentration_sheet(wb)
    elif 'ConcData' in wb.sheetnames:
        return _extract_from_concdata_sheet(wb)
    else:
        wb.close()
        raise ExcelDataExtractionError(
            "Excel file does not contain a 'Concentration' or 'ConcData' sheet. "
            "Expected a Shimadzu NeoBase export file."
        )


def _extract_from_concentration_sheet(wb) -> Tuple[pd.DataFrame, List[str], int]:
    """
    Extract patient data from the 'Concentration' sheet (transposed layout).

    Layout:
    - Row 3: column headers (sample identifiers)
    - Column A (rows 6+): compound names
    - Columns 2-5: criteria limits (skip)
    - BLANK/CONTROL columns: identified by row-3 header containing those words (skip)
    - Remaining columns: patient data (extract)
    """
    ws = wb['Concentration']

    # Step 1: Read row 3 to identify columns
    max_col = ws.max_column
    patient_columns = []  # (col_index, clean_name)

    for c in range(2, max_col + 1):
        header = ws.cell(3, c).value
        if header is None:
            continue

        header_str = str(header).strip()
        if not header_str:
            continue

        upper_header = header_str.upper()

        # Skip criteria limit columns (first few columns with known labels)
        if any(kw in upper_header for kw in [
            'UPPER CAUTION', 'LOWER CAUTION', 'UPPER NOTICE', 'LOWER NOTICE',
            'CAUTION', 'NOTICE'
        ]):
            continue

        # Skip BLANK columns
        if 'BLANK' in upper_header:
            continue

        # Skip CONTROL columns
        if 'CONTROL' in upper_header:
            continue

        # This is a patient column
        patient_columns.append((c, header_str))

    if not patient_columns:
        wb.close()
        raise ExcelDataExtractionError(
            "No patient sample columns found in Concentration sheet row 3."
        )

    # Step 2: Read compound names from column A (rows 6+)
    max_row = ws.max_row
    compound_rows = []  # (row_index, compound_name)

    for r in range(6, max_row + 1):
        compound = ws.cell(r, 1).value
        if compound is None:
            continue
        compound_str = str(compound).strip()
        if compound_str:
            compound_rows.append((r, compound_str))

    if not compound_rows:
        wb.close()
        raise ExcelDataExtractionError(
            "No compound names found in column A of Concentration sheet."
        )

    # Step 3: Extract data — each patient column becomes a row in the output
    compound_names = [name for _, name in compound_rows]
    data_rows = []
    sample_names = []

    for col_idx, raw_name in patient_columns:
        clean_name = _clean_sample_name(raw_name)
        sample_names.append(clean_name)

        row_values = []
        for row_idx, _ in compound_rows:
            val = ws.cell(row_idx, col_idx).value
            if val is None or val == '-' or val == '':
                row_values.append(0.0)
            else:
                try:
                    row_values.append(float(val))
                except (ValueError, TypeError):
                    row_values.append(0.0)

        data_rows.append(row_values)

    wb.close()

    # Step 4: Build DataFrame and reorder columns to match reference_ranges order
    df = pd.DataFrame(data_rows, columns=compound_names)

    from app.core.reference_ranges import range_dict
    expected_order = [c for c in range_dict.keys() if c in df.columns]
    extra_cols = [c for c in df.columns if c not in expected_order]
    df = df[expected_order + extra_cols]

    df.insert(0, 'Sample text', sample_names)

    df = df.apply(lambda col: pd.to_numeric(col, errors='coerce') if col.name != 'Sample text' else col)

    return df, sample_names, len(sample_names)


def _extract_from_concdata_sheet(wb) -> Tuple[pd.DataFrame, List[str], int]:
    """
    Extract data from the 'ConcData' sheet (row-based layout, legacy format).
    Includes controls (duplicated to 4 rows for downstream pipeline).
    """
    ws = wb['ConcData']

    # Step 1: Extract analyte headers from row 6 (odd columns starting at 7)
    analyte_columns = {}  # col_index -> analyte_name
    for c in range(7, ws.max_column + 1, 2):
        header = ws.cell(6, c).value
        if header and isinstance(header, str) and header.strip():
            analyte_columns[c] = header.strip()

    if not analyte_columns:
        wb.close()
        raise ExcelDataExtractionError(
            "No analyte headers found in row 6 of ConcData sheet."
        )

    # Step 2: Find the first data row (skip header/metadata rows 1-11)
    first_data_row = 12
    max_row = ws.max_row

    # Step 3: Extract all sample rows, categorizing as control/blank/patient
    controls = []  # (row_idx, name, is_control_1, is_control_2)
    patients = []  # (row_idx, name)

    for r in range(first_data_row, max_row + 1):
        sample_name = ws.cell(r, 1).value
        if sample_name is None:
            continue

        sample_name = str(sample_name).strip()
        if not sample_name:
            continue

        upper_name = sample_name.upper()

        # Skip blanks entirely
        if upper_name.startswith('BLANK'):
            continue

        # Identify controls
        if 'CONTROL' in upper_name:
            if '1' in upper_name or (len(controls) < 2):
                controls.append((r, sample_name, True, False))
            else:
                controls.append((r, sample_name, False, True))
        else:
            patients.append((r, sample_name))

    # Duplicate controls to match expected 4-row layout
    if len(controls) == 2:
        controls = [controls[0], controls[0], controls[1], controls[1]]
    elif len(controls) < 4:
        print(
            f"Warning: Expected 4 control rows (2x Control I, 2x Control II), "
            f"found {len(controls)}. Proceeding with available controls."
        )

    ordered_rows = controls + patients
    total_count = len(ordered_rows)

    if total_count == 0:
        wb.close()
        raise ExcelDataExtractionError("No valid samples found in Excel file.")

    # Step 4: Extract data for each sample
    compound_names = list(analyte_columns.values())
    data_rows = []
    sample_names = []

    for row_idx, sample_name, *_ in ordered_rows:
        clean_name = _clean_sample_name(sample_name)
        sample_names.append(clean_name)

        row_values = []
        for col_idx in analyte_columns.keys():
            val = ws.cell(row_idx, col_idx).value
            if val is None or val == '-' or val == '':
                row_values.append(0.0)
            else:
                try:
                    row_values.append(float(val))
                except (ValueError, TypeError):
                    row_values.append(0.0)

        data_rows.append(row_values)

    wb.close()

    # Step 5: Build DataFrame
    df = pd.DataFrame(data_rows, columns=compound_names)

    from app.core.reference_ranges import range_dict
    expected_order = [c for c in range_dict.keys() if c in df.columns]
    extra_cols = [c for c in df.columns if c not in expected_order]
    df = df[expected_order + extra_cols]

    df.insert(0, 'Sample text', sample_names)

    df = df.apply(lambda col: pd.to_numeric(col, errors='coerce') if col.name != 'Sample text' else col)

    return df, sample_names, total_count


def extract_date_code_from_filename(filename: str) -> str:
    """
    Extract date code from Excel filename.

    Supports formats like:
    - "18022026 MR LABS DATA.xlsm" -> "18022026"
    - "18022026_MR_LABS.xlsx" -> "18022026"

    Args:
        filename: The Excel filename

    Returns:
        8-digit date code string (DDMMYYYY)

    Raises:
        ExcelDataExtractionError: If no valid date code found
    """
    # Strip path, get just filename
    basename = filename.rsplit('/', 1)[-1].rsplit('\\', 1)[-1]

    # Try to find 8-digit date code at start of filename
    match = re.match(r'^(\d{8})', basename)
    if match:
        date_code = match.group(1)
        # Validate date components
        day = int(date_code[0:2])
        month = int(date_code[2:4])
        year = int(date_code[4:8])

        if 1 <= day <= 31 and 1 <= month <= 12 and 2000 <= year <= 2100:
            return date_code

    raise ExcelDataExtractionError(
        f"Cannot extract date code from filename '{filename}'. "
        f"Expected filename starting with DDMMYYYY (e.g., '18022026 MR LABS DATA.xlsm')."
    )


def _clean_sample_name(raw_name: str) -> str:
    """
    Clean up sample name from Excel format.

    Examples:
    - "718_28_028" -> "718"
    - "12091159_5_005" -> "12091159"
    - "CONTROL 1 _3_003" -> "CONTROL1"
    - "CONTROL 2_4_004" -> "CONTROL2"

    Args:
        raw_name: Raw sample name from Excel

    Returns:
        Cleaned sample name
    """
    name = raw_name.strip()

    # Handle control names
    if 'CONTROL' in name.upper():
        # Extract control number
        ctrl_match = re.match(r'CONTROL\s*(\d+)', name, re.IGNORECASE)
        if ctrl_match:
            return f"CONTROL{ctrl_match.group(1)}"
        return name

    # For patient samples, strip the _XX_YYY suffix
    # Pattern: ID_sequence_run (e.g., "718_28_028")
    parts = name.split('_')
    if len(parts) >= 2:
        # Return just the first part (sample ID)
        return parts[0]

    return name
