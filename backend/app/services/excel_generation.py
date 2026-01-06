# Excel generation module adapted from Vijayrekha_python_reports
# Handles Excel file creation with highlighting based on reference ranges

import os
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
import pandas as pd
from app.core.reference_ranges import range_dict, control_1_range_dict, control_2_range_dict

class ExcelGenerationError(Exception):
    """Custom exception for Excel generation errors"""
    pass

def cell_highlight(min_value: float, max_value: float, cell, bold_font, yellow_fill, green_fill):
    """
    Highlight cells based on their values relative to reference range

    Args:
        min_value: Minimum acceptable value
        max_value: Maximum acceptable value
        cell: Excel cell to highlight
        bold_font: Font object for bold formatting
        yellow_fill: Fill object for out-of-range values
        green_fill: Fill object for in-range values
    """
    if min_value >= max_value:
        raise ExcelGenerationError(f"Invalid range: min ({min_value}) >= max ({max_value})")

    try:
        if not (min_value <= cell.value <= max_value):
            # Out of range - yellow fill
            if cell.value > max_value:
                # Exceeds upper limit - bold font
                cell.font = bold_font
            cell.fill = yellow_fill
        else:
            # In range - green fill
            cell.fill = green_fill
    except Exception as e:
        raise ExcelGenerationError(f"Error highlighting cell: {str(e)}")

def write_df_to_sheet(df: pd.DataFrame, sheet, start_row: int) -> bool:
    """
    Write DataFrame to Excel sheet

    Args:
        df: DataFrame to write
        sheet: Excel worksheet object
        start_row: Starting row number (0-indexed)

    Returns:
        True if successful
    """
    try:
        # Write column headers
        for col_idx, col_name in enumerate(df.columns, start=1):
            sheet.cell(row=start_row + 2, column=col_idx, value=col_name)

        # Write data rows
        for idx, row in df.iterrows():
            for col_idx, value in enumerate(row):
                if pd.isna(value):
                    sheet.cell(row=start_row + idx + 3, column=col_idx + 1, value=None)
                else:
                    sheet.cell(row=start_row + idx + 3, column=col_idx + 1, value=value)

        return True

    except Exception as e:
        raise ExcelGenerationError(f"Error writing DataFrame to sheet: {str(e)}")

def write_to_excel(final_df: list, excel_path: str, date: str, template_path: str) -> str:
    """
    Generate Excel files with highlighting for all patients

    Args:
        final_df: List of 3 DataFrames [Control I, Control II, Patient data]
        excel_path: Path where to save the final_results.xlsx file
        date: Date code from filename (DDMMYYYY)
        template_path: Path to the Excel template file

    Returns:
        Path to the generated final_results.xlsx file

    Raises:
        ExcelGenerationError: If Excel generation fails
    """
    try:
        # Validate template exists
        if not os.path.exists(template_path):
            raise ExcelGenerationError(f"Template file not found: {template_path}")

        # Load template
        wb = load_workbook(template_path)

        # Create Final_Results sheet
        ws = wb.create_sheet(title='Final_Results')

        # DataFrames and their starting rows
        dfs = [(final_df[0], 0), (final_df[1], 10), (final_df[2], 21)]

        # Write all DataFrames to the sheet
        for df, row in dfs:
            if not write_df_to_sheet(df, ws, row):
                raise ExcelGenerationError("Failed to write DataFrame to sheet")

        # Apply cell highlighting
        _apply_cell_highlighting(ws, final_df)

        # Create output directory if needed
        save_dir = _create_output_directory(excel_path, date)

        # Update output path
        excel_path = os.path.join(save_dir, os.path.basename(excel_path))

        # Save main Excel file
        wb.save(excel_path)
        print(f"Excel File for Formatted File: {excel_path}")

        # Generate per-patient Excel files
        _generate_patient_files(final_df[2], excel_path, template_path)

        return excel_path

    except Exception as e:
        raise ExcelGenerationError(f"Error generating Excel files: {str(e)}")

def _apply_cell_highlighting(ws, final_df: list):
    """Apply color highlighting to cells based on reference ranges"""
    # Define styles
    bold = Font(bold=True)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col in final_df[0].columns:
        if col in ["CONTROLS", ""]:
            continue

        if col not in range_dict:
            raise ExcelGenerationError(f"Unknown column '{col}' - no reference range defined")

        col_idx = final_df[0].columns.get_loc(col) + 1
        col_letter = ws.cell(row=1, column=col_idx).column_letter

        row = 1
        for cell in ws[col_letter][1:]:
            if isinstance(cell.value, (int, float)):
                # Determine which range to use based on row position
                if col in control_1_range_dict or col in control_2_range_dict:
                    if row <= 5:
                        if row <= 3:
                            # Control I range
                            if col in control_1_range_dict:
                                min_value, max_value = control_1_range_dict[col]
                                cell_highlight(min_value, max_value, cell, bold, yellow_fill, green_fill)
                    elif row <= 10:
                        if row <= 8:
                            # Control II range
                            if col in control_2_range_dict:
                                min_value, max_value = control_2_range_dict[col]
                                cell_highlight(min_value, max_value, cell, bold, yellow_fill, green_fill)
                    else:
                        # Patient range
                        min_value, max_value = range_dict[col]
                        cell_highlight(min_value, max_value, cell, bold, yellow_fill, green_fill)
                else:
                    if row > 6:
                        # Patient range for compounds not in control ranges
                        min_value, max_value = range_dict[col]
                        cell_highlight(min_value, max_value, cell, bold, yellow_fill, green_fill)

                row += 1

def _create_output_directory(excel_path: str, date: str) -> str:
    """
    Create output directory with incremental naming if directory exists

    Args:
        excel_path: Desired output path
        date: Date code (DDMMYYYY)

    Returns:
        Final output directory path
    """
    i = 1
    save_dir = os.path.dirname(excel_path)

    while True:
        if not os.path.exists(save_dir):
            os.makedirs(save_dir)
            break
        else:
            # Directory exists, check if it matches the date
            dir_parts = save_dir.split(os.sep)
            if dir_parts[-1] == date:
                # Correct date, append increment
                dir_parts[-1] = f"{date}({i})"
            else:
                # Wrong date, replace it
                dir_parts[-1] = f"{date}({i})"

            save_dir = os.sep.join(dir_parts)
            i += 1

    return save_dir

def _generate_patient_files(patient_df: pd.DataFrame, final_results_path: str, template_path: str):
    """
    Generate individual Excel files for each patient

    Args:
        patient_df: DataFrame containing patient data
        final_results_path: Path to final_results.xlsx
        template_path: Path to Excel template
    """
    for index, row in patient_df.iterrows():
        patient_name = row['Sample text']

        # Skip non-patient rows
        if patient_name in ['Reference Range', '']:
            continue

        # Load fresh template for each patient
        temp_wb = load_workbook(final_results_path)

        # Get Sheet1 and populate patient data
        temp_sheet1 = temp_wb['Sheet1']
        temp_sheet1['B3'] = patient_name
        temp_sheet1['K3'] = patient_name

        # Assign values per patient
        start_row = 4
        for idx, value in enumerate(row):
            if isinstance(value, (int, float)):
                temp_sheet1.cell(row=start_row + idx, column=11, value=value)

        # Rename Sheet2 to patient name
        temp_sheet2 = temp_wb['Sheet2']
        temp_sheet2.title = f'{patient_name}'

        # Save patient workbook
        patient_file_path = final_results_path.replace('final_results', f'{patient_name}')
        temp_wb.save(patient_file_path)
        print(f"Excel file created for {patient_name}")
