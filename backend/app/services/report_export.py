"""
Report Export Service - Generate Excel and HTML exports
"""

from pathlib import Path
from typing import Dict, List
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime


def generate_excel_export(analysis_data: Dict, output_path: Path) -> Path:
    """
    Generate Excel export with multiple sheets.

    Args:
        analysis_data: Dictionary containing analysis results
        output_path: Path to save the Excel file

    Returns:
        Path to the generated Excel file
    """
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    abnormal_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
    normal_fill = PatternFill(start_color="E6FFE6", end_color="E6FFE6", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Sheet 1: Summary
    ws_summary = wb.create_sheet("Summary")
    summary = analysis_data['summary']

    ws_summary['A1'] = 'Neonatal Screening Report Analysis'
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:B1')

    ws_summary['A3'] = 'Report File:'
    ws_summary['B3'] = analysis_data['file_name']
    ws_summary['A4'] = 'Analysis Date:'
    ws_summary['B4'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    ws_summary['A6'] = 'Summary Statistics'
    ws_summary['A6'].font = Font(bold=True, size=14)

    ws_summary['A7'] = 'Total Tests:'
    ws_summary['B7'] = summary['total_tests']
    ws_summary['A8'] = 'Normal Results:'
    ws_summary['B8'] = summary['normal_count']
    ws_summary['A9'] = 'Abnormal Results:'
    ws_summary['B9'] = summary['abnormal_count']
    ws_summary['A10'] = 'Status:'
    ws_summary['B10'] = summary['status'].upper()

    # Apply status color
    if summary['status'] == 'abnormal':
        ws_summary['B10'].fill = abnormal_fill
        ws_summary['B10'].font = Font(bold=True, color="C00000")
    else:
        ws_summary['B10'].fill = normal_fill
        ws_summary['B10'].font = Font(bold=True, color="00C000")

    # Adjust column widths
    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 40

    # Sheet 2: Patient Information
    ws_patient = wb.create_sheet("Patient Information")
    ws_patient['A1'] = 'Patient Information'
    ws_patient['A1'].font = Font(bold=True, size=14)

    patient_info = analysis_data['patient_info']
    row = 3
    for key, value in patient_info.items():
        if value:
            ws_patient[f'A{row}'] = key.replace('_', ' ').title() + ':'
            ws_patient[f'A{row}'].font = Font(bold=True)
            ws_patient[f'B{row}'] = value
            row += 1

    ws_patient.column_dimensions['A'].width = 20
    ws_patient.column_dimensions['B'].width = 40

    # Sheet 3: Abnormalities
    if analysis_data['abnormalities']:
        ws_abnormal = wb.create_sheet("Abnormalities")

        # Headers
        headers = ['Category', 'Analyte', 'Value', 'Unit', 'Reference Range', 'Reason']
        for col, header in enumerate(headers, 1):
            cell = ws_abnormal.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border

        # Data
        for row_idx, abn in enumerate(analysis_data['abnormalities'], 2):
            ws_abnormal.cell(row=row_idx, column=1, value=abn['category']).border = border
            ws_abnormal.cell(row=row_idx, column=2, value=abn['analyte']).border = border
            ws_abnormal.cell(row=row_idx, column=3, value=abn['value']).border = border
            ws_abnormal.cell(row=row_idx, column=4, value=abn['unit']).border = border
            ws_abnormal.cell(row=row_idx, column=5, value=abn['reference_range']).border = border
            ws_abnormal.cell(row=row_idx, column=6, value=abn['reason']).border = border

            # Apply abnormal fill to entire row
            for col in range(1, 7):
                ws_abnormal.cell(row=row_idx, column=col).fill = abnormal_fill

        # Adjust column widths
        ws_abnormal.column_dimensions['A'].width = 20
        ws_abnormal.column_dimensions['B'].width = 25
        ws_abnormal.column_dimensions['C'].width = 12
        ws_abnormal.column_dimensions['D'].width = 8
        ws_abnormal.column_dimensions['E'].width = 18
        ws_abnormal.column_dimensions['F'].width = 30

    # Sheet 4: All Test Results
    ws_all = wb.create_sheet("All Test Results")

    # Combine all test results
    all_tests = []

    # Amino Acids
    for test in analysis_data['amino_acids']:
        all_tests.append({
            'Category': 'Amino Acid',
            'Analyte': test['analyte'],
            'Value': test['value'],
            'Unit': test['unit'],
            'Reference Range': test['reference_range'],
            'Status': 'Normal' if test.get('is_normal', True) else 'Abnormal'
        })

    # Amino Acid Ratios
    for test in analysis_data['amino_acid_ratios']:
        all_tests.append({
            'Category': 'Amino Acid Ratio',
            'Analyte': test['ratio'],
            'Value': test['value'],
            'Unit': '',
            'Reference Range': test['reference_range'],
            'Status': 'Normal' if test.get('is_normal', True) else 'Abnormal'
        })

    # Acylcarnitines
    for test in analysis_data['acylcarnitines']:
        all_tests.append({
            'Category': 'Acylcarnitine',
            'Analyte': test['analyte'],
            'Value': test['value'],
            'Unit': test['unit'],
            'Reference Range': test['reference_range'],
            'Status': 'Normal' if test.get('is_normal', True) else 'Abnormal'
        })

    # Acylcarnitine Ratios
    for test in analysis_data['acylcarnitine_ratios']:
        all_tests.append({
            'Category': 'Acylcarnitine Ratio',
            'Analyte': test['ratio'],
            'Value': test['value'],
            'Unit': '',
            'Reference Range': test['reference_range'],
            'Status': 'Normal' if test.get('is_normal', True) else 'Abnormal'
        })

    # Create DataFrame and write to sheet
    df = pd.DataFrame(all_tests)

    # Headers
    for col, header in enumerate(df.columns, 1):
        cell = ws_all.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data
    for row_idx, row_data in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_all.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border

            # Apply color based on status
            if value == 'Abnormal':
                for col in range(1, 7):
                    ws_all.cell(row=row_idx, column=col).fill = abnormal_fill
            elif value == 'Normal':
                for col in range(1, 7):
                    ws_all.cell(row=row_idx, column=col).fill = normal_fill

    # Adjust column widths
    ws_all.column_dimensions['A'].width = 20
    ws_all.column_dimensions['B'].width = 25
    ws_all.column_dimensions['C'].width = 12
    ws_all.column_dimensions['D'].width = 8
    ws_all.column_dimensions['E'].width = 18
    ws_all.column_dimensions['F'].width = 12

    # Save workbook
    wb.save(output_path)
    return output_path


def generate_html_export(analysis_data: Dict) -> str:
    """
    Generate HTML export with professional styling.

    Args:
        analysis_data: Dictionary containing analysis results

    Returns:
        HTML string
    """
    summary = analysis_data['summary']
    patient_info = analysis_data['patient_info']
    abnormalities = analysis_data['abnormalities']

    # Determine status color
    status_color = '#dc2626' if summary['status'] == 'abnormal' else '#10b981'
    status_bg = '#fee2e2' if summary['status'] == 'abnormal' else '#d1fae5'

    html = f"""
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Neonatal Screening Report Analysis</title>
    <style>
        * {{
            margin: 0;
            padding: 0;
            box-sizing: border-box;
        }}

        body {{
            font-family: -apple-system, BlinkMacSystemFont, 'Segoe UI', Roboto, Oxygen, Ubuntu, Cantarell, sans-serif;
            line-height: 1.6;
            color: #2d3748;
            background: #f7fafc;
            padding: 2rem;
        }}

        .container {{
            max-width: 1200px;
            margin: 0 auto;
            background: white;
            border-radius: 12px;
            box-shadow: 0 4px 12px rgba(0, 0, 0, 0.1);
            overflow: hidden;
        }}

        .header {{
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white;
            padding: 2rem;
            text-align: center;
        }}

        .header h1 {{
            font-size: 2rem;
            margin-bottom: 0.5rem;
        }}

        .header p {{
            opacity: 0.9;
        }}

        .content {{
            padding: 2rem;
        }}

        .section {{
            margin-bottom: 2rem;
        }}

        .section h2 {{
            font-size: 1.5rem;
            color: #2d3748;
            margin-bottom: 1rem;
            padding-bottom: 0.5rem;
            border-bottom: 2px solid #e2e8f0;
        }}

        .summary-grid {{
            display: grid;
            grid-template-columns: repeat(4, 1fr);
            gap: 1rem;
            margin-bottom: 2rem;
        }}

        .summary-card {{
            background: #f7fafc;
            border-radius: 8px;
            padding: 1.5rem;
            text-align: center;
            border: 2px solid #e2e8f0;
        }}

        .summary-card .value {{
            font-size: 2.5rem;
            font-weight: 700;
            color: #2d3748;
            margin-bottom: 0.5rem;
        }}

        .summary-card .label {{
            font-size: 0.9rem;
            color: #718096;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}

        .summary-card.status {{
            background: {status_bg};
            border-color: {status_color};
        }}

        .summary-card.status .value {{
            color: {status_color};
        }}

        .info-table {{
            width: 100%;
            border-collapse: collapse;
        }}

        .info-table tr {{
            border-bottom: 1px solid #e2e8f0;
        }}

        .info-table tr:last-child {{
            border-bottom: none;
        }}

        .info-table td {{
            padding: 0.75rem;
        }}

        .info-table td:first-child {{
            font-weight: 600;
            color: #4a5568;
            width: 200px;
        }}

        .abnormality-table {{
            width: 100%;
            border-collapse: collapse;
            margin-top: 1rem;
        }}

        .abnormality-table th {{
            background: #ef4444;
            color: white;
            padding: 0.75rem;
            text-align: left;
            font-weight: 600;
        }}

        .abnormality-table td {{
            padding: 0.75rem;
            border-bottom: 1px solid #fecaca;
        }}

        .abnormality-table tr {{
            background: #fef2f2;
        }}

        .abnormality-table tr:hover {{
            background: #fee2e2;
        }}

        .no-abnormalities {{
            background: #d1fae5;
            border: 2px solid #10b981;
            border-radius: 8px;
            padding: 2rem;
            text-align: center;
            color: #065f46;
        }}

        .no-abnormalities h3 {{
            color: #047857;
            margin-bottom: 0.5rem;
        }}

        .footer {{
            background: #f7fafc;
            padding: 1.5rem 2rem;
            text-align: center;
            color: #718096;
            font-size: 0.9rem;
            border-top: 1px solid #e2e8f0;
        }}

        @media print {{
            body {{
                background: white;
                padding: 0;
            }}

            .container {{
                box-shadow: none;
            }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>Neonatal Screening Report Analysis</h1>
            <p>Comprehensive Analysis Report</p>
        </div>

        <div class="content">
            <!-- Summary Section -->
            <div class="section">
                <h2>Analysis Summary</h2>
                <div class="summary-grid">
                    <div class="summary-card">
                        <div class="value">{summary['total_tests']}</div>
                        <div class="label">Total Tests</div>
                    </div>
                    <div class="summary-card">
                        <div class="value">{summary['normal_count']}</div>
                        <div class="label">Normal</div>
                    </div>
                    <div class="summary-card">
                        <div class="value">{summary['abnormal_count']}</div>
                        <div class="label">Abnormal</div>
                    </div>
                    <div class="summary-card status">
                        <div class="value">{summary['status'].upper()}</div>
                        <div class="label">Status</div>
                    </div>
                </div>
            </div>

            <!-- Patient Information -->
            <div class="section">
                <h2>Patient Information</h2>
                <table class="info-table">
                    <tr>
                        <td>Report File:</td>
                        <td>{analysis_data['file_name']}</td>
                    </tr>
    """

    # Add patient info
    for key, value in patient_info.items():
        if value:
            label = key.replace('_', ' ').title()
            html += f"""
                    <tr>
                        <td>{label}:</td>
                        <td>{value}</td>
                    </tr>
    """

    html += """
                </table>
            </div>
    """

    # Abnormalities Section
    if abnormalities:
        html += """
            <div class="section">
                <h2>Abnormalities Detected</h2>
                <table class="abnormality-table">
                    <thead>
                        <tr>
                            <th>Category</th>
                            <th>Analyte</th>
                            <th>Value</th>
                            <th>Unit</th>
                            <th>Reference Range</th>
                            <th>Reason</th>
                        </tr>
                    </thead>
                    <tbody>
        """

        for abn in abnormalities:
            html += f"""
                        <tr>
                            <td>{abn['category']}</td>
                            <td><strong>{abn['analyte']}</strong></td>
                            <td>{abn['value']}</td>
                            <td>{abn['unit']}</td>
                            <td>{abn['reference_range']}</td>
                            <td>{abn['reason']}</td>
                        </tr>
            """

        html += """
                    </tbody>
                </table>
            </div>
        """
    else:
        html += """
            <div class="section">
                <div class="no-abnormalities">
                    <h3>✓ All Clear!</h3>
                    <p>No abnormalities detected. All test values are within normal range.</p>
                </div>
            </div>
        """

    # Footer
    html += f"""
        </div>

        <div class="footer">
            <p>Report generated on {datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
            <p>Generated by Neonatal Report Analyzer</p>
        </div>
    </div>
</body>
</html>
    """

    return html
