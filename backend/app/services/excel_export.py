# Excel export service for NBS report data with color coding and reference ranges

import io
from typing import Dict, List, Any
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

class ExcelExportError(Exception):
    """Custom exception for Excel export errors"""
    pass

def export_review_data_to_excel(processed_data: Dict[str, Any], edited_data: Dict[str, float] = None) -> io.BytesIO:
    """
    Export the review grid data to Excel with color coding and reference ranges

    Args:
        processed_data: The processed report data with all compounds, ratios, and biochemical params
        edited_data: Optional dictionary of edited cell values {cellKey: value}

    Returns:
        BytesIO object containing the Excel file
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "NBS Report Data"

        # Define color fills to match the frontend
        green_fill = PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")
        red_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
        edited_fill = PatternFill(start_color="FFE5CC", end_color="FFE5CC", fill_type="solid")
        header_fill = PatternFill(start_color="4A90E2", end_color="4A90E2", fill_type="solid")

        # Define fonts
        header_font = Font(bold=True, color="FFFFFF", size=11)
        bold_font = Font(bold=True)

        # Define borders
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Get all compounds (columns)
        compounds = processed_data['compounds']

        # Create header row
        headers = ['Sample Name', 'Type'] + compounds
        ws.append(headers)

        # Style header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        # Add data rows
        row_num = 2
        for idx, row_data in enumerate(processed_data['processed_data']):
            sample_name = row_data['sample_name']

            # Determine row type
            if row_data['is_control_1']:
                row_type = 'Control I'
            elif row_data['is_control_2']:
                row_type = 'Control II'
            else:
                row_type = 'Patient'

            # Write sample name and type
            ws.cell(row=row_num, column=1, value=sample_name)
            ws.cell(row=row_num, column=1).font = bold_font
            ws.cell(row=row_num, column=1).border = thin_border

            ws.cell(row=row_num, column=2, value=row_type)
            ws.cell(row=row_num, column=2).border = thin_border
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal='center')

            # Write compound values
            for col_idx, compound in enumerate(compounds, 3):
                compound_data = row_data['values'].get(compound, {})
                value = compound_data.get('value')
                color = compound_data.get('color', 'none')

                # Check if this cell was edited
                cell_key = f"{idx}-{compound}"
                is_edited = edited_data and cell_key in edited_data
                if is_edited:
                    value = edited_data[cell_key]

                cell = ws.cell(row=row_num, column=col_idx)

                # Set value
                if value is not None:
                    cell.value = round(float(value), 2)
                    cell.number_format = '0.00'
                else:
                    cell.value = '—'

                # Apply color based on threshold
                if is_edited:
                    # Edited cells get a special color overlay
                    cell.fill = edited_fill
                elif color == 'green':
                    cell.fill = green_fill
                elif color == 'yellow':
                    cell.fill = yellow_fill
                elif color == 'red':
                    cell.fill = red_fill
                    cell.font = Font(bold=True, color="721C24")

                cell.alignment = Alignment(horizontal='right')
                cell.border = thin_border

            row_num += 1

        # Add reference ranges section
        row_num += 2
        ws.cell(row=row_num, column=1, value="Reference Ranges")
        ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        row_num += 1

        # Patient ranges header
        ws.cell(row=row_num, column=1, value="Parameter")
        ws.cell(row=row_num, column=2, value="Patient Range")
        ws.cell(row=row_num, column=3, value="Control I Range")
        ws.cell(row=row_num, column=4, value="Control II Range")
        for col in range(1, 5):
            cell = ws.cell(row=row_num, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        row_num += 1

        # Add reference ranges for compounds
        ref_ranges = processed_data['reference_ranges']

        # Patient compound ranges
        for compound in compounds:
            if compound in ref_ranges['patient']:
                min_val, max_val = ref_ranges['patient'][compound]
                ws.cell(row=row_num, column=1, value=compound)
                ws.cell(row=row_num, column=2, value=f"{min_val}-{max_val}")

                # Control I range
                if compound in ref_ranges['control_1']:
                    c1_min, c1_max = ref_ranges['control_1'][compound]
                    ws.cell(row=row_num, column=3, value=f"{c1_min}-{c1_max}")

                # Control II range
                if compound in ref_ranges['control_2']:
                    c2_min, c2_max = ref_ranges['control_2'][compound]
                    ws.cell(row=row_num, column=4, value=f"{c2_min}-{c2_max}")

                for col in range(1, 5):
                    ws.cell(row=row_num, column=col).border = thin_border

                row_num += 1

            # Ratio ranges
            elif compound in ref_ranges.get('ratios', {}):
                min_val, max_val = ref_ranges['ratios'][compound]
                ws.cell(row=row_num, column=1, value=compound)
                ws.cell(row=row_num, column=2, value=f"{min_val}-{max_val}")
                ws.cell(row=row_num, column=3, value="N/A")
                ws.cell(row=row_num, column=4, value="N/A")

                for col in range(1, 5):
                    ws.cell(row=row_num, column=col).border = thin_border

                row_num += 1

            # Biochemical parameter ranges
            elif compound in ref_ranges.get('biochemical', {}):
                min_val, max_val = ref_ranges['biochemical'][compound]
                ws.cell(row=row_num, column=1, value=compound)
                ws.cell(row=row_num, column=2, value=f"{min_val}-{max_val}")
                ws.cell(row=row_num, column=3, value="N/A")
                ws.cell(row=row_num, column=4, value="N/A")

                for col in range(1, 5):
                    ws.cell(row=row_num, column=col).border = thin_border

                row_num += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 25  # Sample Name
        ws.column_dimensions['B'].width = 12  # Type
        for col_idx in range(3, len(compounds) + 3):
            ws.column_dimensions[get_column_letter(col_idx)].width = 12

        # Freeze panes (freeze first row and first two columns)
        ws.freeze_panes = 'C2'

        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return output

    except Exception as e:
        raise ExcelExportError(f"Failed to export Excel: {str(e)}")
